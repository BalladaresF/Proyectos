from openpyxl import Workbook
from openpyxl.styles import Font
import openpyxl
import time

book = Workbook()
sheet = book.active

#manejo de celdas:
sheet['A1']=5
sheet['A2']=10

#manejo de cadena de celdas:
sheet['B1'] = 'Rango'
sheet['B1'].font = Font(color='FF0000', bold=True)
for i in range(2, 15):
    sheet[f'B{i}'] = i**2

#creación de una nueva hoja:
sheet2 = book.create_sheet('Hoja_2')
sheet2['A1']='Esta es otra hoja.'

#insertar fecha actual:
fecha=time.strftime('%x')
sheet2['A2'] = fecha

#unión de celdas:
sheet3 = book.create_sheet('hoja_3')
sheet3.merge_cells('A1:D1')
sheet3['A1'] = 'Prueba de  unión de celdas'
#sheet3.unmerge_cells('A1:D1')

#utilizar fórmulas:
sheet['B16'] = '=SUM(B2:B14)'

#guardar archivo en carpeta
path = ('C:\\Users\Andres B\\Desktop\Visual Studio code\\Python projects\\Excels\\')
book.save(path+'Prueba_Uno.xlsx')

#leer datos del excel:
print("\nDesea leer datos de la primera hoja? \n1.Sí\n2.No")
Respuesta=input()

match Respuesta:
    case "1":
        LeerBook = openpyxl.load_workbook(path+'Prueba_Uno.xlsx')
        sheetLec=LeerBook.active
        a1 = sheetLec['A1']
        a2 = sheetLec['A2']
        print()
        print(a1.value)
        print(a2.value)
    case "2":
         print("Creación finalizada.")
    case _:
        print("Respuesta no válida.")